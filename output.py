"""
output.py  ── Excel / テキスト出力
"""

from __future__ import annotations
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import VENUE_CODE_MAP


# ════════════════════════════════════════════════════════════
# スタイル定数
# ════════════════════════════════════════════════════════════

_H_FILL  = PatternFill("solid", fgColor="1F3864")
_H_FONT  = Font(color="FFFFFF", bold=True)
_THIN    = Side(style="thin", color="CCCCCC")
_BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER  = Alignment(horizontal="center", vertical="center")
_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

_BUY    = PatternFill("solid", fgColor="FFD700")
_WARN   = PatternFill("solid", fgColor="FFB347")
_SKIP   = PatternFill("solid", fgColor="D3D3D3")
_HIT    = PatternFill("solid", fgColor="90EE90")
_FLAG   = PatternFill("solid", fgColor="FF6B6B")
_ST_CLR = PatternFill("solid", fgColor="87CEEB")
_VAL_C  = PatternFill("solid", fgColor="FFFACD")
_NIGHT  = PatternFill("solid", fgColor="E8E8FF")
_MISS   = PatternFill("solid", fgColor="FFCCCC")
_HP_ROW = PatternFill("solid", fgColor="FF8C00")   # 高配当行（ダークオレンジ）
_HP_HDR = PatternFill("solid", fgColor="FF4500")   # 高配当シートヘッダ
_HP_GLD = PatternFill("solid", fgColor="FFD700")   # 高配当フラグ確定行（金）
_HP_PNK = PatternFill("solid", fgColor="FFF0F5")   # 高配当スコア準該当行（薄ピンク）


def _style_header(ws, col_widths: list[float] | None = None) -> None:
    for cell in ws[1]:
        cell.fill      = _H_FILL
        cell.font      = _H_FONT
        cell.alignment = _CENTER_WRAP
        cell.border    = _BORDER
    ws.row_dimensions[1].height = 28
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 13


def _style_picks_rows(ws, df: pd.DataFrame) -> None:
    cols = df.columns.tolist()
    jc   = cols.index("判定") + 1
    fc   = cols.index("イン崩壊") + 1 if "イン崩壊" in cols else None
    nc   = cols.index("ナイター") + 1 if "ナイター" in cols else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        j = str(row[jc-1].value or "")
        for c in row:
            c.border    = _BORDER
            c.alignment = _CENTER
        fill = _BUY if "✅" in j else _WARN if "⚠️" in j else _SKIP
        for c in row:
            c.fill = fill
        if fc and str(row[fc-1].value or "") == "⚠️":
            row[fc-1].fill = _FLAG
            row[fc-1].font = Font(bold=True)
        if nc and "🌙" in str(row[nc-1].value or ""):
            row[nc-1].fill = _NIGHT


def _style_detail_rows(ws, df: pd.DataFrame) -> None:
    cols = df.columns.tolist()
    uf   = cols.index("イン崩壊フラグ") + 1 if "イン崩壊フラグ" in cols else None
    sc   = cols.index("ST狙い目フラグ") + 1 if "ST狙い目フラグ" in cols else None
    ov   = cols.index("過小評価フラグ") + 1 if "過小評価フラグ" in cols else None
    hc   = cols.index("着順")           + 1 if "着順"           in cols else None
    vf   = cols.index("バリューフラグ") + 1 if "バリューフラグ" in cols else None
    hpf  = cols.index("高配当フラグ")   + 1 if "高配当フラグ"   in cols else None  # ★追加

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.border    = _BORDER
            c.alignment = _CENTER
        # 高配当フラグを最優先で色付け
        if hpf and row[hpf-1].value == 1:
            for c in row: c.fill = _HP_ROW
        elif vf and row[vf-1].value == 1:
            for c in row: c.fill = _VAL_C
        if uf and row[uf-1].value == 1:
            for c in row: c.fill = _FLAG
        elif sc and row[sc-1].value == 1:
            for c in row: c.fill = _ST_CLR
        elif ov and row[ov-1].value == 1:
            for c in row: c.fill = _BUY
        if hc and row[hc-1].value == 1:
            row[hc-1].fill = _HIT
            row[hc-1].font = Font(bold=True)


# ════════════════════════════════════════════════════════════
# メイン出力関数
# ════════════════════════════════════════════════════════════

def save_picks_excel(
    picks_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    output_file: str | Path,
) -> None:
    """買い目 + 詳細 + 高配当候補 + 場別シート を1つのExcelに出力"""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # 全場買い目シート
        picks_df.to_excel(writer, sheet_name="📋全場買い目", index=False)
        ws = writer.sheets["📋全場買い目"]
        _style_header(ws)
        _style_picks_rows(ws, picks_df)
        ws.freeze_panes = "A2"

        # ── 💎高配当候補シート（払戻8000以上の可能性が高いレース）─────
        hp_cond1 = (
            picks_df["判定"].str.startswith("✅") &
            (picks_df.get("高配当フラグ", pd.Series(0, index=picks_df.index)) == 1)
        ) if "高配当フラグ" in picks_df.columns else pd.Series(False, index=picks_df.index)
        hp_cond2 = (
            picks_df["判定"].str.startswith("✅") &
            (picks_df.get("高配当スコア", pd.Series(0.0, index=picks_df.index)) >= 0.5) &
            (picks_df.get("バリュースコア", pd.Series(0.0, index=picks_df.index)) >= 1.0)
        ) if "高配当スコア" in picks_df.columns else pd.Series(False, index=picks_df.index)

        hp_df = picks_df[hp_cond1 | hp_cond2].copy()
        if "高配当スコア" in hp_df.columns:
            hp_df = hp_df.sort_values("高配当スコア", ascending=False).reset_index(drop=True)

        hp_show = [c for c in [
            "場名","レースNo","判定","軸艇番","軸選手","軸確率","軸真期待値",
            "軸推定オッズ","高配当スコア","予測払戻","高配当フラグ",
            "バリュースコア","動的EV_MIN","ナイター",
            "3連単①","3連単②","3連単③","注意フラグ","イン崩壊",
        ] if c in hp_df.columns]

        if len(hp_df) > 0:
            hp_df[hp_show].to_excel(writer, sheet_name="💎高配当候補(8000+)", index=False)
            ws_hp = writer.sheets["💎高配当候補(8000+)"]
            # ヘッダー（赤）
            for cell in ws_hp[1]:
                cell.fill      = _HP_HDR
                cell.font      = Font(color="FFFFFF", bold=True)
                cell.alignment = _CENTER_WRAP
                cell.border    = _BORDER
            ws_hp.row_dimensions[1].height = 28
            for col in ws_hp.columns:
                ws_hp.column_dimensions[col[0].column_letter].width = 14
            # 行スタイル（高配当フラグ確定=金、準該当=薄ピンク）
            fl_idx = hp_show.index("高配当フラグ") + 1 if "高配当フラグ" in hp_show else None
            for row in ws_hp.iter_rows(min_row=2, max_row=ws_hp.max_row):
                fl_val = int(row[fl_idx-1].value or 0) if fl_idx else 0
                fill_r = _HP_GLD if fl_val == 1 else _HP_PNK
                for c in row:
                    c.fill      = fill_r
                    c.alignment = _CENTER
                    c.border    = _BORDER
            ws_hp.freeze_panes = "A2"
            print(f"[高配当候補] {len(hp_df)}R → 💎高配当候補(8000+) シートに出力")
        else:
            pd.DataFrame(columns=["メッセージ"]).to_excel(
                writer, sheet_name="💎高配当候補(8000+)", index=False)
            writer.sheets["💎高配当候補(8000+)"]["A2"] = \
                "高配当候補なし（条件を満たすレースがありませんでした）"
            print("[高配当候補] 該当レースなし")

        # 詳細シート
        detail_df.to_excel(writer, sheet_name="🔍選手詳細", index=False)
        ws2 = writer.sheets["🔍選手詳細"]
        _style_header(ws2)
        _style_detail_rows(ws2, detail_df)
        ws2.freeze_panes = "A2"

        # 場ごとシート
        for vc in picks_df["場コード"].unique():
            vn = VENUE_CODE_MAP.get(str(vc).zfill(2), f"場{vc}")
            sn = vn[:8]

            vp = picks_df[picks_df["場コード"] == vc]
            sheet_name = f"{sn}_買"
            vp.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_header(writer.sheets[sheet_name])
            _style_picks_rows(writer.sheets[sheet_name], vp)

            vd = detail_df[detail_df["場名"] == vn]
            if len(vd) > 0:
                sheet_name2 = f"{sn}_詳"
                vd.to_excel(writer, sheet_name=sheet_name2, index=False)
                _style_header(writer.sheets[sheet_name2])
                _style_detail_rows(writer.sheets[sheet_name2], vd)

    print(f"[完了] {output_file}")


def save_result_excel(
    result_df: pd.DataFrame,
    pred_result: pd.DataFrame,
    output_file: str | Path,
) -> None:
    """結果照合Excelを出力"""
    show_r = [
        "場名","レースNo","艇番","選手名","予測_1着確率","真期待値",
        "アンサンブルスコア_IN順位","着順","1着フラグ","払戻",
        "バリュースコア","バリューフラグ","動的EV_MIN","ナイターフラグ",
        "イン崩壊フラグ","過小評価フラグ",
    ]
    show_r = [c for c in show_r if c in pred_result.columns]

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # 的中詳細
        result_df.to_excel(writer, sheet_name="的中詳細", index=False)
        ws = writer.sheets["的中詳細"]
        _style_header(ws)
        if "的中" in result_df.columns:
            hc = result_df.columns.tolist().index("的中") + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for c in row:
                    c.border    = _BORDER
                    c.alignment = _CENTER
                    c.fill      = _HIT if row[hc-1].value else _MISS

        # 予実突合
        (pred_result[show_r]
         .sort_values(["場名","レースNo","アンサンブルスコア_IN順位"])
         .to_excel(writer, sheet_name="予実突合", index=False))
        ws2 = writer.sheets["予実突合"]
        _style_header(ws2)
        if "1着フラグ" in show_r:
            hc2 = show_r.index("1着フラグ") + 1
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                for c in row:
                    c.border    = _BORDER
                    c.alignment = _CENTER
                if row[hc2-1].value == 1:
                    for c in row:
                        c.fill = _HIT

    print(f"[完了] {output_file}")


# ════════════════════════════════════════════════════════════
# コンソール統計出力
# ════════════════════════════════════════════════════════════

def print_summary(
    result_df: pd.DataFrame,
    pred_result: pd.DataFrame,
    total_bet: float,
    total_ret: float,
) -> None:
    n_r = len(result_df)
    n_h = result_df["的中"].sum() if n_r > 0 else 0
    roi = total_ret / total_bet * 100 if total_bet > 0 else 0

    print(f"\n━━ 的中結果 ━━")
    if n_r > 0:
        print(f"  購入: {n_r}R / 的中: {n_h}R ({n_h/n_r*100:.1f}%)")
    else:
        print("  購入なし")
    print(f"  投資: ¥{total_bet:,.0f} / 回収: ¥{total_ret:,.0f} / 回収率: {roi:.1f}%")

    if len(result_df) > 0:
        print("\n━━ 場別成績 ━━")
        for vn, vg in result_df.groupby("場名"):
            vh = vg["的中"].sum()
            vr = vg["回収"].sum()
            vb = len(vg) * 100
            print(f"  {vn}: {len(vg)}R / 的中{vh}R / 回収率{vr/vb*100:.1f}%")

    valid_r = pred_result[pred_result["返還フラグ"] == 0]

    if "バリュースコア" in valid_r.columns:
        print("\n━━ バリュースコア帯別ROI ━━")
        for lo, hi, lb in [
            (0, 1.0, "<1.0"), (1.0, 1.2, "1.0-1.2"), (1.2, 1.5, "1.2-1.5"),
            (1.5, 2.0, "1.5-2.0"), (2.0, 9999, ">2.0"),
        ]:
            seg = valid_r[(valid_r["バリュースコア"] >= lo) & (valid_r["バリュースコア"] < hi)]
            if len(seg) == 0:
                continue
            roi_s = (seg["1着フラグ"].mean() * (seg["払戻"].mean() / 100)) \
                    if seg["払戻"].notna().any() else 0
            print(f"  {lb}: {len(seg)}件 / 1着率={seg['1着フラグ'].mean():.3f} / 理論ROI={roi_s:.2f}")

    print("\n━━ EV帯別ROI ━━")
    for lo, hi, lb in [
        (0, 1.0, "<1.0"), (1.0, 1.5, "1.0-1.5"), (1.5, 2.0, "1.5-2.0"),
        (2.0, 5.0, "2.0-5.0"), (5.0, 9999, ">5.0"),
    ]:
        seg = valid_r[(valid_r["真期待値"] >= lo) & (valid_r["真期待値"] < hi)]
        if len(seg) == 0:
            continue
        roi_s = (seg["1着フラグ"].mean() * (seg["払戻"].mean() / 100)) \
                if seg["払戻"].notna().any() else 0
        print(f"  {lb}: {len(seg)}件 / 1着率={seg['1着フラグ'].mean():.3f} / 理論ROI={roi_s:.2f}")

    if "ナイターフラグ" in valid_r.columns:
        print("\n━━ ナイター vs 昼間 ━━")
        for flag, lb in [(1, "🌙ナイター"), (0, "☀️昼間")]:
            seg = valid_r[valid_r["ナイターフラグ"] == flag]
            if len(seg) > 0:
                print(f"  {lb}: {len(seg)}件 / 1着率={seg['1着フラグ'].mean():.3f}")
